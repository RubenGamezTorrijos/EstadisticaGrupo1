"""
PROYECTO: Estadística para Ingeniería
INTEGRACIÓN Y UI: RUBEN GAMEZ TORRIJOS (Coordinador y Desarrollador)
App principal de Streamlit para visualizar el análisis estadístico.
"""
import streamlit as st
import pandas as pd
import os
from analisis.estadisticos import limpiar_datos, calcular_estadisticos
from analisis.graficos import crear_histograma, crear_boxplot, crear_scatter_regresion, crear_bar_chart, crear_grafico_comparativo_ic
from analisis.inferencial import calcular_ic_95, contraste_hipotesis, verificar_supuestos, verificar_homocedasticidad
import io
from fpdf import FPDF

# Configuración de página RUBEN
st.set_page_config(
    page_title="Estadística - Salarios en IT",
    page_icon="🤖",
    layout="wide"
)

# Diccionario de traducción de variables para visualización
VAR_LABELS = {
    'work_year': 'Año de Trabajo',
    'experience_level': 'Nivel de Experiencia',
    'employment_type': 'Tipo de Empleo',
    'job_title': 'Título del Puesto',
    'salary': 'Salario (Local)',
    'salary_currency': 'Moneda',
    'salary_in_usd': 'Salario (USD)',
    'employee_residence': 'Residencia Empleado',
    'remote_ratio': 'Ratio Remoto',
    'company_location': 'Localización Empresa',
    'company_size': 'Tamaño Empresa',
    'job_category': 'Categoría de Puesto',
    'work_setting': 'Modalidad de Trabajo',
    'cost_of_living_index': 'Índice de Coste de Vida'
}

# Estilo Blue (#0B84F4) Adaptativo
st.markdown(f"""
    <style>
    /* Estilo para las métricas - Adaptativo */
    [data-testid="stMetric"] {{
        background-color: rgba(11, 132, 244, 0.05);
        padding: 20px;
        border-radius: 12px;
        border: 1px solid #0B84F4;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    }}
    
    /* Contenedores de tablas */
    [data-testid="stTable"] {{
        background-color: transparent;
        border-radius: 8px;
    }}
    
    /* Barra lateral */
    [data-testid="stSidebar"] {{
        border-right: 1px solid rgba(11, 132, 244, 0.2);
    }}
    
    /* Títulos y acentos */
    h1, h2, h3 {{ 
        font-weight: 700; 
    }}
    
    /* Configuración de Pestañas (Tabs) - Adaptativo */
    .stTabs [data-baseweb="tab-list"] {{ 
        gap: 10px; 
    }}
    .stTabs [data-baseweb="tab"] {{ 
        height: 45px; 
        background-color: rgba(11, 132, 244, 0.05); 
        border-radius: 8px 8px 0 0; 
        color: inherit;
        padding: 0 20px !important;
        font-weight: 500;
        border: 1px solid transparent;
        transition: all 0.3s;
    }}
    .stTabs [data-baseweb="tab"]:hover {{
        border-color: #0B84F4;
        color: #0B84F4;
    }}
    .stTabs [aria-selected="true"] {{ 
        background-color: #0B84F4 !important; 
        color: white !important;
        border-bottom: 2px solid #0B84F4;
    }}
    
    /* Botones y Selectores */
    .stButton>button {{
        background-color: #0B84F4;
        color: white;
        border-radius: 8px;
        border: none;
        width: 100%;
    }}
    
    .stRadio [data-baseweb="radio"] div[aria-checked="true"] > div:first-child {{
        background-color: #0B84F4 !important;
        border-color: #0B84F4 !important;
    }}

    /* Ajuste para inputs en modo claro */
    .stSelectbox div[data-baseweb="select"] {{
        border-radius: 8px;
    }}
    </style>
""", unsafe_allow_html=True)

# Mapeo de nombres de moneda para mayor claridad (Requerimiento Rubén)
CURRENCY_DISPLAY = {
    'EUR': 'Euro (€)',
    'USD': 'Dólar ($)',
    'GBP': 'Libra (£)',
    'CAD': 'Dólar Can. (CAD)',
    'AUD': 'Dólar Aus. (AUD)',
    'BRL': 'Real (BRL)',
    'PLN': 'Zloty (PLN)',
    'CHF': 'Franco Suizo (CHF)'
}

# Localización de números EUR / USD y Abreviaturas
def abbreviate_number(val):
    if val >= 1_000_000:
        return val / 1_000_000, "M"
    elif val >= 1_000:
        return val / 1_000, "K"
    return val, ""

def fmt_eur(val, is_int=False):
    """Formato Europa: 1.234,56"""
    try:
        val, suffix = abbreviate_number(val)
        if is_int and suffix == "":
            res = f"{val:,.0f}"
        else:
            res = f"{val:,.2f}"
            if res.endswith(".00"):
                res = res[:-3]
        return res.replace(",", "@").replace(".", ",").replace("@", ".") + suffix
    except:
        return val

def fmt_usd(val, is_int=False):
    """Formato USA: 1,234.56"""
    try:
        val, suffix = abbreviate_number(val)
        if is_int and suffix == "":
             res = f"{val:,.0f}"
        else:
             res = f"{val:,.2f}"
             if res.endswith(".00"):
                 res = res[:-3]
        return res + suffix
    except:
        return val

def format_currency(val, var_name, is_int=False):
    """Aplica el formato correcto según la variable."""
    if var_name == 'salary_in_usd':
        return fmt_usd(val, is_int)
    else:
        return fmt_eur(val, is_int)

@st.cache_data
def get_data():
    if not os.path.exists('datos/dataset_crudo.csv'):
        # Crear copia si no existe
        if os.path.exists('datos/jobs_in_data.csv'):
            import shutil
            shutil.copy('datos/jobs_in_data.csv', 'datos/dataset_crudo.csv')
    
    df = pd.read_csv('datos/dataset_crudo.csv')
    return limpiar_datos(df)

def create_excel(df):
    """Generar un reporte Excel con pestañas siguiendo el enunciado_practica.md"""
    output = io.BytesIO()
    
    # Secciones según la práctica
    # 1. Resumen (Escritorio)
    df_resumen = pd.DataFrame([
        {'Concepto': 'Fecha de Análisis', 'Dato': pd.Timestamp.now().strftime('%d/%m/%Y')},
        {'Concepto': 'Dataset', 'Dato': 'Jobs in Data IT'},
        {'Concepto': 'Nº Registros Analizados', 'Dato': len(df)},
        {'Concepto': 'Variables Estudiadas', 'Dato': 'Salario, Año, Experiencia, Categoría'}
    ])

    # 2. Análisis Descriptivo
    df_stats = calcular_estadisticos(df)
    
    # 3. Datos para Inferencia (Ejemplo de IC)
    ic_95 = calcular_ic_95(df['salary_in_usd'])
    if ic_95 and 'media' in ic_95:
        df_inf = pd.DataFrame([
            {'Métrica': 'Media Salarial', 'Valor': ic_95['media']},
            {'Métrica': 'IC 95% Inferior', 'Valor': ic_95['ic_inferior']},
            {'Métrica': 'IC 95% Superior', 'Valor': ic_95['ic_superior']},
            {'Métrica': 'Margen de Error', 'Valor': ic_95['margen_error']}
        ])
    else:
        df_inf = pd.DataFrame([{'Métrica': 'Aviso', 'Valor': 'Datos insuficientes para inferencia'}])

    # 4. Análisis de Regresión (NUEVO)
    # Calculamos regresión básica para el Excel
    fig_r, stats_reg = crear_scatter_regresion(df, 'cost_of_living_index', 'salary_in_usd')
    df_reg = pd.DataFrame([
        {'Métrica': 'Variable X', 'Valor': 'Índice de Coste de Vida'},
        {'Métrica': 'Variable Y', 'Valor': 'Salario (USD)'},
        {'Métrica': 'Coeficiente Correlación (r)', 'Valor': stats_reg['correlacion']},
        {'Métrica': 'Coeficiente Determinación (R²)', 'Valor': stats_reg['r_cuadrado']},
        {'Métrica': 'Pendiente', 'Valor': stats_reg['pendiente']},
        {'Métrica': 'Interpretación', 'Valor': 'Tendencia positiva' if stats_reg['pendiente'] > 0 else 'Tendencia negativa'}
    ])

    # 5. Datos del Equipo
    df_equipo = pd.DataFrame([
        {'Nombre': 'Rafael Rodriguez', 'Rol': 'Data Manager'},
        {'Nombre': 'Bryann Vallejo', 'Rol': 'Analista Inferencial'},
        {'Nombre': 'Leslie Ross', 'Rol': 'Analista Descriptivo'},
        {'Nombre': 'Rubén Gámez', 'Rol': 'Coordinador e Integrador'}
    ])

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, index=False, sheet_name='1. Resumen Ejecutivo')
        df_stats.to_excel(writer, index=False, sheet_name='2. Descriptiva')
        df_inf.to_excel(writer, index=False, sheet_name='3. Inferencia')
        df_reg.to_excel(writer, index=False, sheet_name='4. Regresion')
        df_equipo.to_excel(writer, index=False, sheet_name='5. Equipo')
        df.to_excel(writer, index=False, sheet_name='6. Datos Brutos')

        # --- INSERCIÓN DE GRÁFICOS EN EXCEL (Requerimiento Rubén) ---
        from openpyxl.drawing.image import Image
        import tempfile

        # Gráfico Descriptivo (Histograma)
        fig_hist = crear_histograma(df, 'salary_in_usd')
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            fig_hist.savefig(tmp.name, format='png', bbox_inches='tight')
            img_hist = Image(tmp.name)
            writer.sheets['2. Descriptiva'].add_image(img_hist, 'H2')

        # Gráfico Inferencial (IC)
        fig_inf = crear_grafico_comparativo_ic(df, 'experience_level', 'salary_in_usd')
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            fig_inf.savefig(tmp.name, format='png', bbox_inches='tight')
            img_inf = Image(tmp.name)
            writer.sheets['3. Inferencia'].add_image(img_inf, 'H2')

        # Gráfico Regresión
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            fig_r.savefig(tmp.name, format='png', bbox_inches='tight')
            img_reg = Image(tmp.name)
            writer.sheets['4. Regresion'].add_image(img_reg, 'H2')

    return output.getvalue()

def sanitize_pdf_text(text):
    """Limpia caracteres que rompen la fuente Helvetica estándar de FPDF"""
    if not isinstance(text, str):
        text = str(text)
    replacements = {
        '₀': '0', '₁': '1', '₂': '2', '₃': '3', '₄': '4',
        'µ': 'mu', 'μ': 'mu', 'σ': 'sigma', 'π': 'pi',
        '²': '^2', '±': '+/-', '≥': '>=', '≤': '<=', '≠': '!='
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    # Codificar a latin-1 ignorando errores si es necesario para FPDF estándar
    return text.encode('latin-1', 'replace').decode('latin-1')

def create_pdf(df):
    """Genera un informe técnico completo siguiendo el enunciado_practica.md"""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # --- PÁGINA 1: PORTADA ---
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 24)
    pdf.set_text_color(11, 132, 244)
    pdf.cell(0, 60, sanitize_pdf_text("MEMORIA TÉCNICA"), ln=True, align='C')
    pdf.set_font("helvetica", 'B', 16)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, sanitize_pdf_text("ANÁLISIS ESTADÍSTICO: SALARIOS EN IT"), ln=True, align='C')
    pdf.ln(30)
    
    # Detectar país para el título del PDF
    monedas_unicas = df['salary_currency'].unique()
    ubicacion = "Global" if len(df['company_location'].unique()) > 1 else df['company_location'].iloc[0]
    
    pdf.set_font("helvetica", 'B', 20)
    pdf.set_text_color(11, 132, 244)
    pdf.cell(0, 20, f"Informe Estadistico IT - {ubicacion}", ln=True, align='C')
    pdf.ln(10)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", size=10)
    pdf.cell(0, 7, "Rafael Rodriguez - Data Manager", ln=True, align='C')
    pdf.cell(0, 7, "Bryann Vallejo - Analista Inferencial", ln=True, align='C')
    pdf.cell(0, 7, "Leslie Ross - Analista Descriptivo", ln=True, align='C')
    pdf.cell(0, 7, "Ruben Gamez - Coordinador e Integrador", ln=True, align='C')
    pdf.ln(20)
    pdf.set_font("helvetica", 'I', 11)
    pdf.cell(0, 10, "Universidad Europea - Grupo 1", ln=True, align='C')
    
    # --- PÁGINA 2: ANÁLISIS DESCRIPTIVO ---
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 16)
    pdf.set_text_color(11, 132, 244)
    pdf.cell(0, 10, "1. Analisis Descriptivo", ln=True)
    pdf.ln(5)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", size=11)
    stats_df = calcular_estadisticos(df)
    
    # Búsqueda robusta por ID_Variable para evitar errores de indexación
    try:
        salario_stats = stats_df[stats_df['ID_Variable'] == 'salary_in_usd'].iloc[0]
    except (IndexError, KeyError):
        if not stats_df.empty:
            salario_stats = stats_df.iloc[0]
        else:
            # Si no hay datos, generar una página de aviso
            pdf.add_page()
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, "Aviso: Datos insuficientes", ln=True)
            pdf.set_font("helvetica", size=12)
            pdf.multi_cell(0, 10, "No se han encontrado registros suficientes para generar el análisis detallado para el filtro seleccionado.")
            return bytes(pdf.output())

    
    desc_text = (
        f"Se han analizado {len(df)} registros de salarios. "
        f"La media salarial es de ${salario_stats['Media']:,.2f} USD, "
        f"con una desviacion tipica de {salario_stats['Desviación Típica']:,.2f}. "
        f"La mediana se situa en ${salario_stats['Mediana']:,.2f} USD."
    )
    pdf.multi_cell(0, 7, sanitize_pdf_text(desc_text))
    pdf.ln(5)
    
    # Gráfico 1: Histograma
    fig_hist = crear_histograma(df, 'salary_in_usd')
    img_hist = io.BytesIO()
    fig_hist.savefig(img_hist, format='png', bbox_inches='tight', dpi=100)
    pdf.image(img_hist, x=15, w=180)
    pdf.ln(5)
    
    # Gráfico 2: Boxplot (NUEVO)
    fig_box = crear_boxplot(df, 'salary_in_usd', 'experience_level')
    img_box = io.BytesIO()
    fig_box.savefig(img_box, format='png', bbox_inches='tight', dpi=100)
    pdf.ln(5)
    pdf.image(img_box, x=15, w=180)
    
    # --- PÁGINA 3: ESTADÍSTICA INFERENCIAL ---
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 16)
    pdf.set_text_color(11, 132, 244)
    pdf.cell(0, 10, "2. Estadistica Inferencial", ln=True)
    pdf.ln(5)
    
    # Intervalo de Confianza con protección
    ic_95 = calcular_ic_95(df['salary_in_usd'])
    if ic_95 and 'media' in ic_95:
        inf_text = (
            "Aplicando la distribucion T de Student (confianza 95%), estimamos que la media "
            f"poblacional se encuentra entre ${ic_95['ic_inferior']:,.2f} y ${ic_95['ic_superior']:,.2f} USD. "
            f"El margen de error es de +/- ${ic_95['margen_error']:,.2f} USD."
        )
        pdf.multi_cell(0, 7, sanitize_pdf_text(inf_text))
        
        # Gráfico de apoyo inferencial (NUEVO)
        fig_ic = crear_grafico_comparativo_ic(df, 'experience_level', 'salary_in_usd')
        img_ic = io.BytesIO()
        fig_ic.savefig(img_ic, format='png', bbox_inches='tight', dpi=100)
        pdf.image(img_ic, x=15, w=180)
        pdf.ln(5)
    else:
        pdf.set_text_color(255, 0, 0)
        pdf.cell(0, 10, "Aviso: Muestra insuficiente para calculos inferenciales.", ln=True)
        pdf.set_text_color(0, 0, 0)
    
    pdf.ln(10)
    
    # Contraste de Hipótesis con protección
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "Contraste de Hipotesis (Senior vs Mid-level)", ln=True)
    
    sal_senior = df[df['experience_level'] == 'Senior']['salary_in_usd']
    sal_mid = df[df['experience_level'] == 'Mid-level']['salary_in_usd']
    
    if len(sal_senior) > 1 and len(sal_mid) > 1:
        res_test = contraste_hipotesis(sal_senior, sal_mid, "Senior", "Mid-level")
        pdf.set_font("helvetica", size=11)
        test_text = (
            f"Hipotesis Nula (H0): Las medias son iguales.\n"
            f"P-Valor: {res_test['P-Valor']:.4f}\n"
            f"Decision: {res_test['Decisión']}\n"
            f"Conclusion: {res_test['Conclusión']}"
        )
        pdf.multi_cell(0, 7, sanitize_pdf_text(test_text))
    else:
        pdf.set_font("helvetica", size=10)
        pdf.cell(0, 10, "No hay suficientes datos comparativos entre Senior y Mid-level en esta zona.", ln=True)
    
    # --- PÁGINA 4: REGRESIÓN LINEAL ---
    pdf.add_page()
    pdf.set_font("helvetica", 'B', 16)
    pdf.set_text_color(11, 132, 244)
    pdf.cell(0, 10, "3. Relacion y Regresion", ln=True)
    pdf.ln(5)
    
    fig_reg, stats_reg = crear_scatter_regresion(df, 'work_year', 'salary_in_usd')
    reg_text = (
        "Se ha analizado la evolucion temporal de los salarios. "
        f"Se obtuvo una correlacion r = {stats_reg['correlacion']:.4f} "
        f"y un coeficiente de determinacion R2 = {stats_reg['r_cuadrado']:.4f}. "
        f"La pendiente de la recta es {stats_reg['pendiente']:.2f}."
    )
    pdf.set_font("helvetica", size=11)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(0, 7, sanitize_pdf_text(reg_text))
    pdf.ln(5)
    
    img_reg = io.BytesIO()
    fig_reg.savefig(img_reg, format='png', bbox_inches='tight', dpi=100)
    pdf.image(img_reg, x=15, w=180)

    return bytes(pdf.output())

def main():
    # Header de la barra lateral (Arquitectura Rubén)
    st.sidebar.markdown(f"""
        <div style="text-align: center; padding: 10px; border-radius: 50%; background-color: rgba(11, 132, 244, 0.1); border: 2px solid #0B84F4; width: 60px; height: 60px; margin: 0 auto 15px auto; display: flex; align-items: center; justify-content: center;">
            <span style="font-size: 30px;">📊</span>
        </div>
        <h2 style="text-align: center; margin-bottom: 25px;">Estadística y Optimización <br> Grupo 1</h2>
    """, unsafe_allow_html=True)
    st.sidebar.markdown("---")
    menu = st.sidebar.radio(
        "Seleccionar opción:",
        ["Escritorio General", "Análisis Descriptivo", "Estadística Inferencial", "Sobre el Equipo"]
    )

    df = get_data()

    # --- FILTRO GEOGRÁFICO GLOBAL (REQUERIMIENTO RUBÉN) ---
    st.sidebar.markdown("---")
    st.sidebar.subheader("🌐 Filtro Geográfico")
    paises_disp = sorted(df['company_location'].unique().tolist())
    selected_countries = st.sidebar.multiselect(
        "Seleccione Países (Sede):", 
        options=paises_disp, 
        default=None,
        help="Si no selecciona ningún país, se mostrarán resultados para TODO el mundo."
    )
    
    if selected_countries:
        df = df[df['company_location'].isin(selected_countries)].copy()
        sufijo_titulo = f" - {', '.join(selected_countries)}"
    else:
        sufijo_titulo = " - Global"

    # --- DETECCIÓN DINÁMICA DE MONEDA (REQUERIMIENTO RUBÉN) ---
    monedas_unicas = df['salary_currency'].unique()
    if len(monedas_unicas) == 1:
        m_code = monedas_unicas[0]
        m_display = CURRENCY_DISPLAY.get(m_code, m_code)
        VAR_LABELS['salary'] = f"Salario ({m_display})"
        moneda_actual = m_code
    else:
        # Si hay varias, listamos las 3 principales o usamos un término más claro que "Local"
        m_list = ", ".join(list(monedas_unicas)[:3])
        VAR_LABELS['salary'] = f"Salario (Divisas: {m_list})"
        moneda_actual = "Varias"

    # --- SECCIÓN DE EXPORTACIÓN (ARQUITECTURA RUBÉN) ---
    st.sidebar.markdown("---")
    st.sidebar.subheader("📥 Exportar Datos")
    
    col1, col2 = st.sidebar.columns(2)
    
    with col1:
        excel_data = create_excel(df)
        st.download_button(
            label="📊 Excel",
            data=excel_data,
            file_name="dataset_salarios_it.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    with col2:
        try:
            pdf_data = create_pdf(df)
            st.download_button(
                label="📕 PDF",
                data=pdf_data,
                file_name="resumen_estadistico.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Error PDF: {str(e)}")

    if menu == "Escritorio General":
        st.title(f"🚀 Escritorio de Salarios en IT{sufijo_titulo}")
        st.markdown("Bienvenido al centro de control del análisis estadístico de salarios en Data Science.")
        
        # Implementación de métricas principales (Rafael Rodriguez)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Registros", f"{len(df):,}")
        
        # Mostrar media en USD siempre
        m2.metric("Media (USD)", f"${df['salary_in_usd'].mean():,.0f}")
        
        # Mostrar media en moneda local si hay una sola
        if len(monedas_unicas) == 1:
            m3.metric(f"Media ({moneda_actual})", f"{df['salary'].mean():,.0f} {moneda_actual}")
        else:
            m3.metric("Año más reciente", df['work_year'].max())
            
        m4.metric("Categorías Puesto", df['job_category'].nunique())

        st.markdown("---")
        st.markdown("### 📋 Vista Previa de Datos (Rafael Rodriguez)")
        
        # Vista previa de datos con etiquetas legibles
        df_preview = df.head(10).copy()
        df_preview.columns = [VAR_LABELS.get(col, col) for col in df_preview.columns]
        st.dataframe(df_preview, use_container_width=True)

    elif menu == "Análisis Descriptivo":
        st.title("📈 Análisis Descriptivo")
        st.write("Exploración detallada de la distribución y relaciones entre variables.")

        tab1, tab2, tab3 = st.tabs(["📊 Estadísticos", "🖼️ Visualizaciones", "📉 Regresión"])
        
        with tab1:
            st.subheader("Medidas Descriptivas (Rafael Rodriguez)")
            df_stats = calcular_estadisticos(df)
            if not df_stats.empty:
                # Estilos condicionales por fila
                df_format = df_stats.copy()
                cols_to_format = ['Media', 'Mediana', 'Moda', 'Rango', 'Desviación Típica', 'Varianza']
                
                # Usar ID_Variable para el formato y mapeo de etiquetas
                def formato_fila(row):
                    if row['ID_Variable'] == 'salary_in_usd':
                        fmt_func = fmt_usd
                    else:
                        fmt_func = fmt_eur
                    
                    return pd.Series([
                        fmt_func(row['Media']), fmt_func(row['Mediana']),
                        fmt_func(row['Moda']), fmt_func(row['Rango']),
                        fmt_func(row['Desviación Típica']), fmt_func(row['Varianza'])
                    ], index=['Media', 'Mediana', 'Moda', 'Rango', 'Desviación Típica', 'Varianza'])

                df_format[cols_to_format] = df_format.apply(formato_fila, axis=1)
                
                # Sincronizar etiqueta de la variable usando el diccionario dinámico de app.py
                df_format['Variable'] = df_format['ID_Variable'].map(VAR_LABELS).fillna(df_format['Variable'])
                
                # Quitar ID_Variable antes de mostrar al usuario
                df_format = df_format.drop(columns=['ID_Variable'])
                
                st.table(df_format)
            
        with tab2:
            st.subheader("Histogramas y Boxplots (Leslie Ross)")
            col_var, col_cat = st.columns(2)
            with col_var:
                var_sel = st.selectbox("Seleccione Variable Numérica:", ['salary_in_usd', 'salary'], format_func=lambda x: VAR_LABELS.get(x, x))
            with col_cat:
                cat_sel = st.selectbox("Dividir por Categoría:", ['experience_level', 'job_category', 'work_setting', 'company_location'], format_func=lambda x: VAR_LABELS.get(x, x))
            
            c1, c2 = st.columns(2)
            with c1:
                st.pyplot(crear_histograma(df, var_sel))
            with c2:
                st.pyplot(crear_boxplot(df, var_sel, cat_sel))

        with tab3:
            st.subheader("Relación y Regresión (Leslie Ross)")
            st.markdown("Analizando factores predictivos del salario.")
            
            sel_x = st.selectbox("Seleccione Variable Predictora (X):", ['work_year', 'cost_of_living_index'], format_func=lambda x: VAR_LABELS.get(x, x))
            fig_reg, stats_reg = crear_scatter_regresion(df, sel_x, 'salary_in_usd')
            st.pyplot(fig_reg)
            
            # Mostrar métricas de regresión debajo
            c1, c2, c3 = st.columns(3)
            c1.metric("Correlación (r)", f"{stats_reg['correlacion']:.4f}")
            c2.metric("R²", f"{stats_reg['r_cuadrado']:.4f}")
            c3.metric("Pendiente", f"{stats_reg['pendiente']:.2f}")

    elif menu == "Estadística Inferencial":
        st.title("🔬 Análisis Inferencial (Bryann Vallejo)")
        st.write("Deducciones estadísticas sobre la población total de profesionales.")

        st.subheader("1. Intervalos de Confianza (95%)")
        var_ic = st.selectbox("Variable para IC:", ['salary_in_usd', 'salary'], format_func=lambda x: VAR_LABELS.get(x, x))
        
        datos_ic = df[var_ic].dropna()
        if len(datos_ic) > 1:
            ic = calcular_ic_95(datos_ic)
            moneda = "USD" if var_ic == 'salary_in_usd' else ""
            
            c_ic1, c_ic2 = st.columns([2, 1])
            with c_ic1:
                st.success(f"Deducción para **{VAR_LABELS[var_ic]}**:")
                st.info(f"👉 **[{format_currency(ic['ic_inferior'], var_ic)}  —  {format_currency(ic['ic_superior'], var_ic)}] {moneda}**")
            with c_ic2:
                st.metric("Margen de Error", f"± {format_currency(ic['margen_error'], var_ic)} {moneda}")
                st.metric("Media Muestral", f"{format_currency(ic['media'], var_ic)} {moneda}")
        else:
            st.warning("⚠️ No hay suficientes datos para calcular el intervalo de confianza (se requieren al menos 2 registros).")

        st.markdown("---")
        st.subheader("2. Contraste de Hipótesis")
        st.write("**Pregunta:** ¿Existe diferencia significativa entre salarios Senior y Mid-level?")
        
        sal_senior = df[df['experience_level'] == 'Senior']['salary_in_usd']
        sal_mid = df[df['experience_level'] == 'Mid-level']['salary_in_usd']
        
        # Diagnósticos previos
        c_dg1, c_dg2 = st.columns(2)
        with c_dg1:
            norm = verificar_supuestos(df['salary_in_usd'])
            st.write(f"**Test Normalidad ({norm['Prueba']}):**")
            st.code(f"Stat={norm['estadistico']:.4f}, p={norm['p_valor']:.4e}")
            st.caption(norm['Conclusión'])
        with c_dg2:
            homo = verificar_homocedasticidad(sal_senior, sal_mid)
            st.write("**Test Homocedasticidad (Levene):**")
            st.code(f"Stat={homo['estadistico']:.4f}, p={homo['p_valor']:.4e}")
            st.caption(homo['Conclusión'])

        st.info("💡 **Justificación Teórica:** Dado que el tamaño de muestra es elevado (>1.000), el **Teorema del Límite Central (TLC)** permite usar tests paramétricos incluso si la normalidad no es perfecta, ya que la distribución de la media muestral tiende a la normalidad.")

        res_test = contraste_hipotesis(sal_senior, sal_mid, "Nivel Senior", "Nivel Mid")
        
        c1, c2 = st.columns(2)
        with c1:
            p_val = res_test['P-Valor']
            p_display = f"{p_val:.4f}".replace(".", ",") if p_val >= 0.0001 else "< 0,0001"
            st.metric("P-Valor Obtenido", p_display)
        with c2:
            st.metric("Decisión Estadística", res_test['Decisión'])
            
        st.warning(f"**Conclusión:** {res_test['Conclusión']}")
        
        with st.expander("📚 Explicación de Conceptos"):
            st.markdown("""
            *   **P-Valor:** Probabilidad de observar los datos si la hipótesis nula fuera cierta. Si p < 0,05, rechazamos H0.
            *   **Tamaño del Efecto (Cohen's d):** Indica la magnitud de la diferencia. Valores > 0,8 sugieren un efecto grande.
            *   **Teorema del Límite Central:** Fundamental en inferencia. Asegura que la media de una muestra grande sigue una distribución normal independientemente de la forma de la población original.
            """)

    elif menu == "Sobre el Equipo":
        st.title("👥 Equipo de Desarrollo")
        st.markdown("""
        Esta aplicación es el resultado del trabajo colaborativo de 4 ingenieros, siguiendo los requerimientos de la asignatura de Estadística.
        
        | Integrante | Rol | Archivos Desarrollados |
        | :--- | :--- | :--- |
        | **Rafael Rodriguez** | Data Manager | `analisis/estadisticos.py`, archivos CSV en `datos/` |
        | **Bryann Vallejo** | Analista Inferencial | `analisis/inferencial.py`, tablas CSV en `outputs/tablas/` |
        | **Leslie Ross** | Vis. Experta | `analisis/graficos.py`, `generate_plots.py`, gráficos PNG en `outputs/graficos/` |
        | **Ruben Gamez** | Desarrollador y Coordinador | `app.py`, `setup_data.py`, `requirements.txt`, `.gitignore` |
        """)
        
        # Espaciado extra para que el recuadro no esté pegado (Requerimiento Rubén)
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        st.info("💡 **Nota del Coordinador:** El proyecto cumple con todos los requisitos: muestra > 100, variables continuas/discretas/categóricas y análisis bivariable.")

if __name__ == "__main__":
    main()
